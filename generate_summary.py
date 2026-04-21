#!/usr/bin/env python3
"""Generate 2026-04-21 Final Incremental Summary.xlsx"""

from pathlib import Path
import warnings
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE       = Path("/home/user/INCREMENTAL-REVENUE")
PREV       = "2026-04-14"
CURR       = "2026-04-21"
MTS_NEW    = BASE / CURR  / f"{CURR} MTS Inventory Report.xlsm"
AGING_NEW  = BASE / CURR  / f"{CURR} Aging inventory analysis.xlsx"
MTS_PRIOR  = BASE / PREV  / f"{PREV} MTS Inventory Incremental Changes.xlsx"
AGING_PRIOR= BASE / PREV  / f"{PREV} Aging Inventory Incremental Changes.xlsx"
OUTPUT     = BASE / CURR  / f"{CURR} Final Incremental Summary.xlsx"

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

FILL_DARK   = fill("1F3864")
FILL_SUB    = fill("EBF3FB")
FILL_BLUE   = fill("2E75B6")
FILL_COMBO  = fill("F2F2F2")
FILL_INC    = fill("E8F5E9")
FILL_DEC    = fill("FFF0EE")
FILL_NEW    = fill("DDEEFF")
FILL_REM    = fill("F5F5F5")
FILL_NONE   = fill("FFFFFF")

FMT_DOLLAR  = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT     = '0.0%;(0.0%);"-"'

def font(bold=False, size=10, color="000000", white=False):
    return Font(name="Arial", bold=bold, size=size,
                color="FFFFFF" if white else color)

def status_fill(s):
    return {"Increased": FILL_INC, "Decreased": FILL_DEC,
            "New Account": FILL_NEW, "Removed": FILL_REM}.get(s, FILL_NONE)

def cell(ws, row, col, value=None, bold=False, size=10, white=False,
         f=None, num_fmt=None, halign=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, white=white)
    if f:
        c.fill = f
    if num_fmt:
        c.number_format = num_fmt
    if halign:
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    else:
        c.alignment = Alignment(vertical="center", wrap_text=True)
    return c

def merge_row(ws, row, col1, col2, value, bold=False, size=10, white=False,
              f=None, halign="left"):
    cell(ws, row, col1, value, bold=bold, size=size, white=white, f=f, halign=halign)
    ws.merge_cells(start_row=row, start_column=col1,
                   end_row=row, end_column=col2)
    for c in range(col1+1, col2+1):
        ws.cell(row=row, column=c).fill = f or FILL_NONE

# ── Load prior-week totals ────────────────────────────────────────────────────
def prior_totals():
    wb = openpyxl.load_workbook(str(MTS_PRIOR), read_only=True, data_only=True)
    mts_prev_total = list(wb["Executive Summary"].iter_rows(values_only=True))[5][2]
    wb.close()
    wb2 = openpyxl.load_workbook(str(AGING_PRIOR), read_only=True, data_only=True)
    aging_prev_total = list(wb2["Executive Summary"].iter_rows(values_only=True))[5][2]
    wb2.close()
    return float(mts_prev_total), float(aging_prev_total)

# ── MTS data ──────────────────────────────────────────────────────────────────
def load_mts():
    # Country code → name mapping
    allowed = pd.read_excel(str(MTS_NEW), sheet_name="MTS kits allowed", engine="openpyxl")
    cmap = dict(zip(allowed["Country code"], allowed["Country"]))
    cmap["-USCEM"] = "USCEM"

    # Current week raw
    raw = pd.read_excel(str(MTS_NEW), sheet_name="MTS", engine="openpyxl")
    raw["total stock on hand"] = pd.to_numeric(raw["total stock on hand"], errors="coerce").fillna(0)
    curr_total = raw["total stock on hand"].sum()

    # Current week by account
    curr = (raw.dropna(subset=["Account"])
               .groupby(["Country", "Account"], as_index=False)
               .agg(soh_c=("total stock on hand", "sum"),
                    owner=("Procedural Solutions Kit: Owner Name", "first")))

    # Prior week by account
    prev_raw = pd.read_excel(str(MTS_PRIOR), sheet_name="By Account", header=2, engine="openpyxl")
    prev = prev_raw[prev_raw["Country"] != "GRAND TOTAL"].dropna(subset=["Account"]).copy()
    prev = prev[["Country", "Account", "Owner Name", "SOH 2026-04-14 ($)"]].copy()
    prev.columns = ["Country", "Account", "owner_p", "soh_p"]
    prev["soh_p"] = pd.to_numeric(prev["soh_p"], errors="coerce").fillna(0)

    # Merge
    merged = pd.merge(
        prev.rename(columns={"Country": "Country_p", "owner_p": "owner_p"}),
        curr.rename(columns={"Country": "Country_c", "owner": "owner_c"}),
        on="Account", how="outer", indicator=True
    )
    merged["soh_p"] = merged["soh_p"].fillna(0)
    merged["soh_c"] = merged["soh_c"].fillna(0)
    merged["Country"] = merged["Country_p"].fillna(merged["Country_c"])
    merged["Owner Name"] = merged["owner_p"].fillna(merged["owner_c"])
    merged["change"] = merged["soh_c"] - merged["soh_p"]
    merged["pct"] = np.where(merged["soh_p"] != 0, merged["change"] / merged["soh_p"], np.nan)

    def status(r):
        if r["_merge"] == "left_only":  return "New Account"
        if r["_merge"] == "right_only": return "Removed"
        if abs(r["change"]) <= 1e-9:    return "No Change"
        return "Increased" if r["change"] > 0 else "Decreased"

    merged["Status"] = merged.apply(status, axis=1)
    merged = (merged[~((merged["soh_p"] == 0) & (merged["soh_c"] == 0))]
                .sort_values(["Country", "Account"]).reset_index(drop=True))

    # By country
    by_c = (merged.groupby("Country")
                  .agg(soh_p=("soh_p", "sum"), soh_c=("soh_c", "sum"))
                  .reset_index())
    by_c["change"] = by_c["soh_c"] - by_c["soh_p"]
    by_c["pct"]    = np.where(by_c["soh_p"] != 0, by_c["change"] / by_c["soh_p"], np.nan)
    by_c["Status"] = by_c["change"].apply(
        lambda x: "No Change" if abs(x) <= 1e-9 else ("Increased" if x > 0 else "Decreased"))
    by_c = by_c.sort_values("change", ascending=False).reset_index(drop=True)

    return merged, by_c, curr_total, cmap

# ── Aging data ────────────────────────────────────────────────────────────────
def load_aging():
    # Current week
    raw = pd.read_excel(str(AGING_NEW), sheet_name="Sheet1", header=3, engine="openpyxl")
    raw = raw.drop(columns=["Unnamed: 15"], errors="ignore")
    raw["Total_n"] = pd.to_numeric(raw["Total"], errors="coerce").fillna(0)
    raw["Account"] = raw["Account"].fillna("(blank)").astype(str)
    raw["Countries"] = raw["Countries"].fillna("").astype(str)
    curr_total = raw["Total_n"].sum()

    curr = (raw.groupby(["Countries", "Account"], as_index=False)
               .agg(tot_c=("Total_n", "sum"),
                    ko=("Procedural Solutions Kit: Owner Name", "first"),
                    sr=("Sales Rep Manager", "first")))
    curr.columns = ["Country", "Account", "tot_c", "Kit Owner", "Sales Rep Manager"]

    # Prior week
    prev_raw = pd.read_excel(str(AGING_PRIOR), sheet_name="By Account", header=2, engine="openpyxl")
    prev = prev_raw[prev_raw["Country"] != "GRAND TOTAL"].dropna(subset=["Account"]).copy()
    prev = prev[["Country", "Account", "Kit Owner", "Sales Rep Manager", "Total $ 2026-04-14"]].copy()
    prev.columns = ["Country", "Account", "ko_p", "sr_p", "tot_p"]
    prev["tot_p"] = pd.to_numeric(prev["tot_p"], errors="coerce").fillna(0)
    prev["Account"] = prev["Account"].fillna("(blank)").astype(str)

    # Merge on (Country, Account)
    merged = pd.merge(
        prev.rename(columns={"ko_p": "ko_p", "sr_p": "sr_p"}),
        curr.rename(columns={"Country": "Country_c", "Kit Owner": "ko_c",
                              "Sales Rep Manager": "sr_c"}),
        on="Account", how="outer", indicator=True,
        suffixes=("_p", "_c")
    )
    # Resolve Country
    if "Country_p" in merged.columns and "Country_c" in merged.columns:
        merged["Country"] = merged["Country_p"].fillna(merged["Country_c"])
    elif "Country_p" in merged.columns:
        merged["Country"] = merged["Country_p"]
    else:
        merged["Country"] = merged["Country_c"]

    merged["tot_p"] = merged["tot_p"].fillna(0)
    merged["tot_c"] = merged["tot_c"].fillna(0)
    merged["Kit Owner"]        = merged["ko_p"].fillna(merged["ko_c"])
    merged["Sales Rep Manager"]= merged["sr_p"].fillna(merged["sr_c"])
    merged["change"] = merged["tot_c"] - merged["tot_p"]
    merged["pct"]    = np.where(merged["tot_p"] != 0, merged["change"] / merged["tot_p"], np.nan)

    def status(r):
        if r["_merge"] == "left_only":  return "New Account"
        if r["_merge"] == "right_only": return "Removed"
        if abs(r["change"]) <= 1e-9:    return "No Change"
        return "Increased" if r["change"] > 0 else "Decreased"

    merged["Status"] = merged.apply(status, axis=1)
    merged = merged.sort_values(["Country", "Account"]).reset_index(drop=True)

    # By country
    by_c = (merged.groupby("Country")
                  .agg(tot_p=("tot_p", "sum"), tot_c=("tot_c", "sum"))
                  .reset_index())
    by_c["change"] = by_c["tot_c"] - by_c["tot_p"]
    by_c["pct"]    = np.where(by_c["tot_p"] != 0, by_c["change"] / by_c["tot_p"], np.nan)
    by_c["Status"] = by_c["change"].apply(
        lambda x: "No Change" if abs(x) <= 1e-9 else ("Increased" if x > 0 else "Decreased"))
    by_c = by_c.sort_values("change", ascending=False).reset_index(drop=True)

    return merged, by_c, curr_total

# ── Sheet 1: Executive Summary ────────────────────────────────────────────────
def write_exec(ws, mts_df, aging_df, mts_curr_total, aging_curr_total,
               mts_prev_total, aging_prev_total):
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 14

    # Row 1 — title
    ws.row_dimensions[1].height = 25.5
    merge_row(ws, 1, 1, 5,
              "FY26 Q4 — FINAL INCREMENTAL INVENTORY SUMMARY",
              bold=True, size=16, white=True, f=FILL_DARK)

    # Row 2 — subtitle
    ws.row_dimensions[2].height = 13.5
    merge_row(ws, 2, 1, 5,
              f"Consolidated: MTS Inventory  +  Aging Inventory   |   Period: {PREV}  →  {CURR}",
              size=9, f=FILL_SUB)

    # Row 3 — spacer
    ws.row_dimensions[3].height = 7.5

    # Row 4 — section header
    ws.row_dimensions[4].height = 18
    merge_row(ws, 4, 1, 5, "COMBINED KEY METRICS",
              bold=True, size=11, white=True, f=FILL_BLUE)

    # Row 5 — column headers
    ws.row_dimensions[5].height = 27.75
    for col, hdr in enumerate(["Report", f"{PREV} ($)", f"{CURR} ($)", "Change ($)", "% Change"], 1):
        cell(ws, 5, col, hdr, bold=True, white=True, f=FILL_BLUE, halign="center")

    # Rows 6-8 — key metrics
    combined_prev = mts_prev_total + aging_prev_total
    combined_curr = mts_curr_total + aging_curr_total

    rows = [
        ("MTS Inventory (Stock on Hand)", mts_prev_total, mts_curr_total),
        ("Aging Inventory",               aging_prev_total, aging_curr_total),
        ("COMBINED TOTAL",                combined_prev,    combined_curr),
    ]
    for i, (label, prev_val, curr_val) in enumerate(rows, 6):
        ws.row_dimensions[i].height = 15
        chg = curr_val - prev_val
        pct = chg / prev_val if prev_val else None
        f_row = FILL_COMBO if label == "COMBINED TOTAL" else FILL_NONE
        bold_row = label == "COMBINED TOTAL"
        cell(ws, i, 1, label,    bold=bold_row, f=f_row)
        cell(ws, i, 2, prev_val, bold=bold_row, f=f_row, num_fmt=FMT_DOLLAR)
        cell(ws, i, 3, curr_val, bold=bold_row, f=f_row, num_fmt=FMT_DOLLAR)
        cell(ws, i, 4, chg,      bold=True,     f=f_row, num_fmt=FMT_DOLLAR)
        cell(ws, i, 5, pct,      bold=bold_row, f=f_row, num_fmt=FMT_PCT)

    # Row 9 — spacer
    ws.row_dimensions[9].height = 7.5

    # Status sections helper
    def write_status_section(start_row, label, status_counts):
        ws.row_dimensions[start_row].height = 18
        merge_row(ws, start_row, 1, 5, label,
                  bold=True, size=11, white=True, f=FILL_BLUE)
        ws.row_dimensions[start_row + 1].height = 27.75
        cell(ws, start_row+1, 1, "Status",     bold=True, white=True, f=FILL_BLUE, halign="center")
        cell(ws, start_row+1, 2, "# Accounts", bold=True, white=True, f=FILL_BLUE, halign="center")
        for j, s in enumerate(["Increased", "Decreased", "New Account", "Removed", "No Change"]):
            r = start_row + 2 + j
            ws.row_dimensions[r].height = 15
            sf = status_fill(s)
            cell(ws, r, 1, s,                        bold=True, f=sf)
            cell(ws, r, 2, int(status_counts.get(s, 0)), f=sf, halign="center")

    mts_counts   = mts_df["Status"].value_counts()
    aging_counts = aging_df["Status"].value_counts()
    write_status_section(10, "ACCOUNT STATUS — MTS INVENTORY",   mts_counts)
    ws.row_dimensions[17].height = 7.5
    write_status_section(18, "ACCOUNT STATUS — AGING INVENTORY", aging_counts)

# ── Sheet 2: By Country ───────────────────────────────────────────────────────
def write_by_country(ws, mts_by_c, aging_by_c):
    # Column widths
    widths = [10, 18, 18, 15, 10, 12, 3, 12, 18, 18, 15, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — title
    ws.row_dimensions[1].height = 25.5
    merge_row(ws, 1, 1, 13,
              "INCREMENTAL CHANGES BY COUNTRY — COMBINED VIEW",
              bold=True, size=13, white=True, f=FILL_DARK)

    # Row 2 — subtitle
    ws.row_dimensions[2].height = 13.5
    merge_row(ws, 2, 1, 13,
              f"Left: MTS Inventory   |   Right: Aging Inventory   |   {PREV} → {CURR}",
              size=9, f=FILL_SUB)

    # Row 3 — section headers
    ws.row_dimensions[3].height = 18
    merge_row(ws, 3, 1, 6, "MTS INVENTORY — BY COUNTRY",   bold=True, white=True, f=FILL_BLUE)
    ws.cell(row=3, column=7).fill = FILL_NONE
    merge_row(ws, 3, 8, 13, "AGING INVENTORY — BY COUNTRY", bold=True, white=True, f=FILL_BLUE)

    # Row 4 — column headers
    ws.row_dimensions[4].height = 27.75
    mts_hdrs   = ["Country", f"SOH {PREV} ($)", f"SOH {CURR} ($)", "Change ($)", "% Change", "Status"]
    aging_hdrs = ["Country", f"Total {PREV} ($)", f"Total {CURR} ($)", "Change ($)", "% Change", "Status"]
    for i, h in enumerate(mts_hdrs, 1):
        cell(ws, 4, i, h, bold=True, white=True, f=FILL_BLUE, halign="center")
    ws.cell(row=4, column=7).fill = FILL_NONE
    for i, h in enumerate(aging_hdrs, 8):
        cell(ws, 4, i, h, bold=True, white=True, f=FILL_BLUE, halign="center")

    # Data rows
    n_mts   = len(mts_by_c)
    n_aging = len(aging_by_c)
    n_rows  = max(n_mts, n_aging)

    for i in range(n_rows):
        r = i + 5
        ws.row_dimensions[r].height = 15
        ws.cell(row=r, column=7).fill = FILL_NONE

        if i < n_mts:
            row = mts_by_c.iloc[i]
            sf = status_fill(row["Status"])
            cell(ws, r, 1, row["Country"],  bold=True, f=sf)
            cell(ws, r, 2, row["soh_p"],    f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 3, row["soh_c"],    f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 4, row["change"],   bold=True, f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 5, row["pct"],      f=sf, num_fmt=FMT_PCT)
            cell(ws, r, 6, row["Status"],   bold=True, f=sf)

        if i < n_aging:
            row = aging_by_c.iloc[i]
            sf = status_fill(row["Status"])
            cell(ws, r, 8,  row["Country"], bold=True, f=sf)
            cell(ws, r, 9,  row["tot_p"],   f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 10, row["tot_c"],   f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 11, row["change"],  bold=True, f=sf, num_fmt=FMT_DOLLAR)
            cell(ws, r, 12, row["pct"],     f=sf, num_fmt=FMT_PCT)
            cell(ws, r, 13, row["Status"],  bold=True, f=sf)

    # Grand Total rows
    mts_gt_row   = n_rows + 5
    aging_gt_row = n_rows + 5
    ws.row_dimensions[mts_gt_row].height = 15

    for col in range(1, 14):
        ws.cell(row=mts_gt_row, column=col).fill = FILL_DARK

    cell(ws, mts_gt_row, 1, "GRAND TOTAL", bold=True, white=True, f=FILL_DARK)
    cell(ws, mts_gt_row, 2, f"=SUM(B5:B{n_mts+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, mts_gt_row, 3, f"=SUM(C5:C{n_mts+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, mts_gt_row, 4, f"=SUM(D5:D{n_mts+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)
    ws.cell(row=mts_gt_row, column=7).fill = FILL_NONE

    cell(ws, aging_gt_row, 8,  "GRAND TOTAL", bold=True, white=True, f=FILL_DARK)
    cell(ws, aging_gt_row, 9,  f"=SUM(I5:I{n_aging+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, aging_gt_row, 10, f"=SUM(J5:J{n_aging+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, aging_gt_row, 11, f"=SUM(K5:K{n_aging+4})", bold=True, white=True,
         f=FILL_DARK, num_fmt=FMT_DOLLAR)

# ── Sheet 3: MTS By Account ───────────────────────────────────────────────────
def write_mts_by_account(ws, mts_df):
    widths = [8, 42, 26, 18, 18, 17, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1
    ws.row_dimensions[1].height = 25.5
    merge_row(ws, 1, 1, 8, "MTS INVENTORY — INCREMENTAL CHANGES BY ACCOUNT",
              bold=True, size=13, white=True, f=FILL_DARK)

    # Row 2
    ws.row_dimensions[2].height = 13.5
    merge_row(ws, 2, 1, 8, f"Stock on Hand: {PREV}  →  {CURR}",
              size=9, f=FILL_SUB)

    # Row 3 — headers
    ws.row_dimensions[3].height = 27.75
    hdrs = ["Country", "Account", "Owner Name",
            f"SOH {PREV} ($)", f"SOH {CURR} ($)", "Change ($)", "% Change", "Status"]
    for i, h in enumerate(hdrs, 1):
        cell(ws, 3, i, h, bold=True, white=True, f=FILL_BLUE, halign="center")

    # Data rows
    for idx, row in mts_df.iterrows():
        r = idx + 4
        ws.row_dimensions[r].height = 15
        sf = status_fill(row["Status"])
        cell(ws, r, 1, row["Country"],     bold=True, f=sf)
        cell(ws, r, 2, row["Account"],     f=sf)
        cell(ws, r, 3, row["Owner Name"],  f=sf)
        cell(ws, r, 4, row["soh_p"],       f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 5, row["soh_c"],       f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 6, row["change"],      bold=True, f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 7, row["pct"],         f=sf, num_fmt=FMT_PCT)
        cell(ws, r, 8, row["Status"],      bold=True, f=sf)

    # Grand Total
    n = len(mts_df)
    gt = n + 4
    ws.row_dimensions[gt].height = 15
    ws.merge_cells(start_row=gt, start_column=1, end_row=gt, end_column=3)
    for c in range(1, 9):
        ws.cell(row=gt, column=c).fill = FILL_DARK
    cell(ws, gt, 1, "GRAND TOTAL",          bold=True, white=True, f=FILL_DARK)
    cell(ws, gt, 4, f"=SUM(D4:D{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, gt, 5, f"=SUM(E4:E{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, gt, 6, f"=SUM(F4:F{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)

# ── Sheet 4: Aging By Account ─────────────────────────────────────────────────
def write_aging_by_account(ws, aging_df):
    widths = [12, 42, 26, 36, 18, 18, 17, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1
    ws.row_dimensions[1].height = 25.5
    merge_row(ws, 1, 1, 9, "AGING INVENTORY — INCREMENTAL CHANGES BY ACCOUNT",
              bold=True, size=13, white=True, f=FILL_DARK)

    # Row 2
    ws.row_dimensions[2].height = 13.5
    merge_row(ws, 2, 1, 9, f"Comparison: {PREV}  →  {CURR}",
              size=9, f=FILL_SUB)

    # Row 3 — headers
    ws.row_dimensions[3].height = 27.75
    hdrs = ["Country", "Account", "Kit Owner", "Sales Rep Manager",
            f"Total $ {PREV}", f"Total $ {CURR}", "Change ($)", "% Change", "Status"]
    for i, h in enumerate(hdrs, 1):
        cell(ws, 3, i, h, bold=True, white=True, f=FILL_BLUE, halign="center")

    # Data rows
    for idx, row in aging_df.iterrows():
        r = idx + 4
        ws.row_dimensions[r].height = 15
        sf = status_fill(row["Status"])
        cell(ws, r, 1, row["Country"],            bold=True, f=sf)
        cell(ws, r, 2, row["Account"],            f=sf)
        cell(ws, r, 3, row["Kit Owner"],          f=sf)
        cell(ws, r, 4, row["Sales Rep Manager"],  f=sf)
        cell(ws, r, 5, row["tot_p"],              f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 6, row["tot_c"],              f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 7, row["change"],             bold=True, f=sf, num_fmt=FMT_DOLLAR)
        cell(ws, r, 8, row["pct"],                f=sf, num_fmt=FMT_PCT)
        cell(ws, r, 9, row["Status"],             bold=True, f=sf)

    # Grand Total
    n = len(aging_df)
    gt = n + 4
    ws.row_dimensions[gt].height = 15
    ws.merge_cells(start_row=gt, start_column=1, end_row=gt, end_column=4)
    for c in range(1, 10):
        ws.cell(row=gt, column=c).fill = FILL_DARK
    cell(ws, gt, 1, "GRAND TOTAL",          bold=True, white=True, f=FILL_DARK)
    cell(ws, gt, 5, f"=SUM(E4:E{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, gt, 6, f"=SUM(F4:F{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)
    cell(ws, gt, 7, f"=SUM(G4:G{n+3})",     bold=True, white=True, f=FILL_DARK, num_fmt=FMT_DOLLAR)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading data...")
    mts_prev_total, aging_prev_total = prior_totals()
    mts_df, mts_by_c, mts_curr_total, _cmap = load_mts()
    aging_df, aging_by_c, aging_curr_total   = load_aging()

    print(f"  MTS   prior={mts_prev_total:,.2f}  curr={mts_curr_total:,.2f}")
    print(f"  Aging prior={aging_prev_total:,.2f}  curr={aging_curr_total:,.2f}")
    print(f"  MTS accounts: {len(mts_df)}  |  Aging accounts: {len(aging_df)}")

    print("Building workbook...")
    wb = openpyxl.Workbook()
    wb.active.title = "Executive Summary"
    for name in ["By Country", "MTS — By Account", "Aging — By Account"]:
        wb.create_sheet(name)

    write_exec(wb["Executive Summary"],
               mts_df, aging_df,
               mts_curr_total, aging_curr_total,
               mts_prev_total, aging_prev_total)

    write_by_country(wb["By Country"], mts_by_c, aging_by_c)
    write_mts_by_account(wb["MTS — By Account"], mts_df)
    write_aging_by_account(wb["Aging — By Account"], aging_df)

    wb.save(str(OUTPUT))
    print(f"Saved: {OUTPUT}")

if __name__ == "__main__":
    main()
